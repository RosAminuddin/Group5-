import openpyxl
import os
from datetime import datetime
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from step1_scrap import scrap_data


# ---------------------------------
# BASE DIRECTORY (AUTO-DETECT FILE LOCATION)
# ---------------------------------
BASE_DIR = os.path.dirname(os.path.abspath(__file__))

OUTPUT_FOLDER = os.path.join(BASE_DIR, "output")
EXCEL_FILE = os.path.join(OUTPUT_FOLDER, "pandemic_data.xlsx")

HEADER = [
    "ID",
    "COUNTRY",
    "TODAY_CASES",
    "TODAY_DEATHS",
    "TOTAL_CASES",
    "CREATED_AT",
]


def save_to_excel(data: dict) -> str:
    print(data)

    # Create output folder
    if not os.path.exists(OUTPUT_FOLDER):
        os.makedirs(OUTPUT_FOLDER)

    # Create Excel file if not exist
    if not os.path.exists(EXCEL_FILE):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Pandemic Data"

        ws.append(HEADER)

        # ---------- HEADER STYLE ----------
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(fill_type="solid", start_color="4F81BD")
        header_align = Alignment(horizontal="center", vertical="center")
        header_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        for col in range(1, len(HEADER) + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = header_border
            ws.column_dimensions[cell.column_letter].width = 20

        wb.save(EXCEL_FILE)

    # Load Excel file
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active

    # Auto ID (based on row count)
    next_id = ws.max_row

    # Append data row
    ws.append([
        next_id,
        data["country"],
        data["today_cases"],
        data["today_deaths"],
        data["total_cases"],
        datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    ])

    wb.save(EXCEL_FILE)
    return EXCEL_FILE


# -------------------
# MAIN PROGRAM (TEST)
# -------------------
if __name__ == "__main__":
    pandemic = scrap_data("Malaysia")
    file_path = save_to_excel(pandemic)
    print(f"✅ Pandemic data saved to {file_path}")